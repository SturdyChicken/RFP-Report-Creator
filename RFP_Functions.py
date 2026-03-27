import csv  # This library lets us read and reqrite to excel csv files
import os  # Used to check for files (for errors)
import sys  # Used to exit the program (for errors)
from datetime import date, datetime  # ... gives the current date and time ...
from dateutil.relativedelta import (
    relativedelta,
)  # Allows us to compare dates inside the library dataset
from dateutil import parser
from openpyxl import Workbook  # Allows us to write directly to an excel spreadsheet
from openpyxl.styles import (
    PatternFill,
    Font,
    Alignment,
)  # Allows us to more finly adjust the formatting of the excel spreadsheet
import requests  # Allows us to scrape web pages
from bs4 import (
    BeautifulSoup,
)  # Allows us to organise the scraped web data into a more easy to work with format
import re  # Let's us use more robust string searching (case sensitivity, etc)
from transformers import pipeline
import logging  # Lets us disable errors
from transformers import (
    logging as hf_logging,
)  # these last two are just for the next 3 lines of code and make the terminal look cleaner lol
import asyncio  # Allows us to open up a webpage automatically, so we can read addition loaded javascript data without backend api stuff
from playwright.async_api import (
    async_playwright,
)  # This allows async to open, close, and run in backgroun more automatically and smoothly

# specialized code to silence the "tied weights" warnings
hf_logging.set_verbosity_error()
logging.getLogger("transformers").setLevel(logging.ERROR)

# Define a global variable to hold the AI model in memory
_shared_classifier = None


async def async_princeton(url, filename="princeton_grants.html"):
    async with async_playwright() as p:
        # CHANGE 1: Headless=False lets you SEE what is happening and bypasses some bot detection
        browser = await p.chromium.launch(headless=True)

        # Create a context with a real user agent (helps avoid being blocked)
        context = await browser.new_context(
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:140.0) Gecko/20100101 Firefox/140.0"
        )
        page = await context.new_page()

        print(f"Navigating to {url}...")

        # CHANGE 2: 'networkidle' ensures we wait for the background JSON data to arrive
        await page.goto(url, wait_until="networkidle")

        print("Waiting for data to populate...")
        try:
            await page.wait_for_selector(
                "#search-result-count", state="visible", timeout=30000
            )
            print("   Data detected!")

            # Optional: A small safety sleep to ensure all cards render after the header appears
            import asyncio

            await asyncio.sleep(2)

        except Exception as e:
            print(
                f"   Warning: Timed out waiting for data. Saving whatever we have... Error: {e}"
            )

        content = await page.content()
        with open(filename, "w", encoding="utf-8") as f:
            f.write(content)

        print(f"HTML saved to '{filename}' in program folder. Closing browser.")
        await browser.close()
        return filename


def clean_and_deduplicate(rfp_library):
    cleaned_library = []
    seen_fingerprints = set()

    for entry in rfp_library:
        # First, we standardize the deadlines to the same format
        raw_deadline = str(entry.get("Deadline", "")).strip()

        try:
            # Parse the date (handles both "1/1/2026" and "01/01/2026")
            dt_object = datetime.strptime(raw_deadline, "%m/%d/%Y")
            # Re-format it to strictly ensure zero-padding (e.g. "01/01/2026")
            entry["Deadline"] = dt_object.strftime("%m/%d/%Y")
        except ValueError:
            # If it's rolling, don't change anything
            entry["Deadline"] = raw_deadline

        # We use (Name, URL, Deadline) as the unique identity of each rant to check against duplicates
        fingerprint = (
            entry.get("Grant Name", "").strip().lower(),
            entry.get("Deadline", "").strip(),
        )

        if fingerprint in seen_fingerprints:
            continue  # Skip this entry, it is a duplicate

        # If new, add to our list and record the fingerprint
        seen_fingerprints.add(fingerprint)
        cleaned_library.append(entry)

    return cleaned_library


def get_keywords(text, top_k=3):

    global _shared_classifier

    possible_keywords = [
        "Climate & Environment",
        "Data Science",
        "Food & Agriculture",
        "Life Sciences/Health",
        "AI/ Machine Learning",
        "Neuroscience",
        "Physical Sciences",
        "Science/Technology",
        "Public Health",
        "Business & Economics",
        "Law",
        "Economics",
        "Journalism",
        "Justice",
        "Policy",
        "Social Sciences",
        "Arts and Humanities",
        "Museums & Libraries",
        "BIPOC",
        "Education",
        "Sustainability",
        "Public Safety",
        "Conservation",
    ]
    # Load the model weights (Only happens the first time this function runs)
    if _shared_classifier is None:
        # Silence the library's built-in progress bars and warnings
        hf_logging.set_verbosity_error()
        try:
            _shared_classifier = pipeline(
                "zero-shot-classification", model="valhalla/distilbart-mnli-12-1"
            )
        except Exception as e:
            print(f"\nModel load failed: {e}")
            return "Manual Review Needed"
        # Actual Analysis
    try:
        # Ensure text is valid
        if not text or not isinstance(text, str):
            return "Manual Review Needed"
        result = _shared_classifier(text, possible_keywords, multi_label=True)

        sorted_keywords = sorted(
            zip(result["labels"], result["scores"]), key=lambda x: x[1], reverse=True
        )
        top_list = [label for label, score in sorted_keywords[:top_k]]
        return ", ".join(top_list)
    except Exception as e:
        print(f"\n[!] Analysis Failed on text: '{text[:30]}...' | Error: {e}")
        return "Manual Review Needed"


def UCB_csv_reader(file_name):
    ## ---------- FUNCTION OBJECTIVES ------------
    # This file will execute the following:
    # Read the contents of the provided CSV file
    # Create a dictionary to store the data
    # The following data will be exported in the library
    # 0. Granting Organization
    # 1. Grant Name
    # 2. Deadline
    # 3. Funding Amount
    # 4. Keywords
    # 5. URL

    # If a file with the name "file_name" is not found in this script's folder, display error and exit the program
    if not os.path.exists(file_name):
        print(f"ERROR: The file {file_name} was not found")
        print(
            "Please ensure the file has that exact name, and is in the same folder as this script"
        )
        sys.exit()

    # If the program is still running, it means the file was found
    print(f"File found: '{file_name}'")
    print("Processing data now...")

    # Creates an empty list that we can write data into later
    rfp_library = []

    try:
        # "with open" basically opens the file AND gaurentees it will be closed later
        # 'r' is for 'read mode', 'utf-8-sig' is standard for excel csv
        with open(file_name, mode="r", encoding="utf-8-sig") as UCB_csv:

            # Both UCB file and DictReader function use the first column of values as headers, so no problems here
            reader = csv.reader(UCB_csv)

            # We are defining our own column headers, so skip the first column
            next(reader, None)

            for row in reader:
                # Python will crash if it tries to read a file without enough columns
                if len(row) <= 8:
                    continue

                original_dealine = row[2]
                if not original_dealine.strip():
                    # If no dealine is given, set dealine to "rolling"
                    final_dealine = "rolling"
                else:
                    final_dealine = original_dealine

                data_entry = {
                    "Granting Organization": row[0],  # Column A
                    "Grant Name": row[1],  # Column B
                    "Deadline": final_dealine,  # C
                    "Funding Amount": row[3],  # Column D
                    "Keywords": row[6],  # Column G
                    "URL": row[8],  # Column I
                }

                rfp_library.append(data_entry)

            print(
                f"Successfully extracted {len(rfp_library)} RFPs from UCB csv provided"
            )

    except Exception as error:
        # If the file is corrupted, it will show this instead of crashing
        print(f"An unexpected error occurred while reading the file: {error}")

    return rfp_library


def parse_date(date_str):
    ## ---------- FUNCTION OBJECTIVES ------------
    # Attempts to turn a string like '1/15/2026' into a date object
    # Returns "NONE" if false
    if not date_str or date_str.lower() == "rolling":
        return None

    formats = ["%m/%d/%Y", "%Y-%m-%d", "%m-%d-%Y"]

    for fmt in formats:
        try:
            # .date() strips the time info, leaving just the Year-Month-Day
            return datetime.strptime(date_str, fmt).date()
        except ValueError:
            continue
    return None


def report_maker(rfp_library, output_folder):
    ## ---------- FUNCTION OBJECTIVES ------------
    # This file will create a formatted excel rfp report as a csv file with the following specifications
    # ouput into the given folder with the file name informed by current date and time
    # formatted in a readable and professional manor using openpyxl library
    # The following columns:
    # 1. Deadline
    # 2. Funding Organization Name
    # 3. <=HYPERLINK(I#,J#)>
    # 4. Funding Amount
    # 5. Key Words (Description / Abstract?)
    # 6. Natural Science?
    # 7. Social Science?
    # 8. Humanities?
    # 9. Direct URL
    # 10. Grant Name

    ## -------------------- SET UP FILE NAME AND PATH --------------------

    def get_sort_key(entry):
        date_str = str(entry.get("Deadline", ""))

        try:
            # Attempt to turn string "01/07/2026" into a real Date Object
            date_obj = datetime.strptime(date_str, "%m/%d/%Y")

            # If successful, it gets Priority 0
            return (0, date_obj, entry["Granting Organization"])

        except (ValueError, TypeError):
            # If it fails (e.g. "Rolling"), it gets Priority 1
            # We use datetime.max as a placeholder to keep the tuple format consistent
            return (1, datetime.max, entry["Granting Organization"])

    # Apply the sort to the library in place
    rfp_library.sort(key=get_sort_key)

    # gets current time of running the program
    today = date.today()

    ## Start date is 1st date of next month from running
    start_date = (today + relativedelta(months=1)).replace(day=1)

    ## End date is 1st date of month 6 months from running
    end_date = start_date + relativedelta(months=6)

    # string from time, %B is the full month, and %Y is the 4 digit year
    file_name = f"{start_date.strftime('%B_%Y')}_RFP_-_The_College.xlsx"

    if not os.path.exists(output_folder):
        os.makedirs(output_folder)

    # os.path.join works with both windows and mac
    full_path = os.path.join(output_folder, file_name)

    ## -------------------- "BUCKETS" FOR FORMATTING" --------------------

    fill_data_row = PatternFill(
        start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"
    )  # CURRENT COLOR: LIGHT GREY
    fill_month_header = PatternFill(
        start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"
    )  # CURRENT COLOR: LIGHT BLUE
    fill_main_header = PatternFill(
        start_color="8EA9DB", end_color="8EA9DB", fill_type="solid"
    )  # CURRENT COLOR: DAKRER BLUE
    fill_white = PatternFill(
        start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"
    )  # CURRENT COLOR: PURE WHITE
    fill_ns = PatternFill(
        start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"
    )  # CURRENT COLOR: LIGHT ORANGE
    fill_ss = PatternFill(
        start_color="DDEBF7", end_color="DDEBF7", fill_type="solid"
    )  # CURRENT COLOR: LIGHT BLUE
    fill_h = PatternFill(
        start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"
    )  # CURRENT COLOR: LIGHT GREEN
    font_bold = Font(bold=True)
    font_link = Font(color="0000FF", underline="single")  # Blue & Underlined
    align_center = Alignment(horizontal="center", vertical="center")

    align_general = Alignment(vertical="center", wrap_text=False)

    # Now we need to sort the data into "buckets", one for each month + one for rolling
    buckets = {}

    # The first day of the month is what we will sort against
    current_month_cursor = start_date
    while current_month_cursor < end_date:
        buckets[current_month_cursor] = []
        current_month_cursor += relativedelta(months=1)

    # Add a separate bucket for Rolling
    buckets["Rolling"] = []

    # Sort the data into these buckets
    for rfp in rfp_library:
        original_deadline = rfp["Deadline"]

        if original_deadline == "rolling":
            buckets["Rolling"].append(rfp)
        else:
            parsed_date = parse_date(original_deadline)
            if parsed_date and start_date <= parsed_date < end_date:
                # We force all dates to be the first of the month they are in (for the purposes of bucketing)
                bucket_key = parsed_date.replace(day=1)
                if bucket_key in buckets:
                    buckets[bucket_key].append(rfp)
                else:
                    # This is in case a date slips through the filter
                    pass
            # Dates outside the window are not placed in a bucket, so will not be written into the report file

    ## -------------------- SET UP WORKBOOK AND HEADERS --------------------

    # Create the Excel Workbook and Worksheet with NO GRIDLINES
    wb = Workbook()
    ws = wb.active
    ws.title = "RFP Report"
    ws.sheet_view.showGridLines = False

    # Approx conversion: Width = (Pixels - 5) / 7
    ws.column_dimensions["A"].width = 11  # 75px
    ws.column_dimensions["B"].width = 60  # 420px
    ws.column_dimensions["C"].width = 82  # 575px
    ws.column_dimensions["D"].width = 28  # 200px
    ws.column_dimensions["E"].width = 42  # 300px
    ws.column_dimensions["F"].width = 17  # 110px
    ws.column_dimensions["G"].width = 15  # 100px
    ws.column_dimensions["H"].width = 12  # 90px
    ws.column_dimensions["I"].width = 11  # 75px
    ws.column_dimensions["J"].width = 14  # 100px

    # Define Headers
    headers = [
        "Deadline",  # A
        "Granting Organization",  # B
        "Link",  # C
        "Funding Amount",  # D
        "Keywords",  # E
        "Discipline: NS",  # F
        "Discipline: SS",  # G
        "Discipline: H",  # H
        "URL",  # I
        "Grant Name",  # J
    ]

    ws.append(headers)
    ws.row_dimensions[1].height = 25  # Set Header Row Height

    # Add style to main header
    for cell in ws[1]:
        # If it is Column K (11), make it white/invisible. Otherwise, make it Blue.
        if cell.column == 11:
            cell.fill = fill_white
        else:
            cell.fill = fill_main_header
            cell.font = font_bold
            cell.alignment = align_center

    ## -------------------- WRITE BUCKETS TO EXCEL --------------------

    # since we may be skipping items, we need to maintain our own excel counters
    excel_row_index = 2

    # Function-seption!
    def write_rfp_list(rfp_list, block_title):
        # Allows us to modify the index from outside the function, a kind of memory
        nonlocal excel_row_index

        # For block title: Merge cells from Column 1 (A) to 10 (J)
        ws.merge_cells(
            start_row=excel_row_index,
            start_column=1,
            end_row=excel_row_index,
            end_column=10,
        )

        # Set value and style on the first cell (Top-Left of merge)
        title_cell = ws.cell(row=excel_row_index, column=1, value=block_title)
        title_cell.fill = fill_month_header
        title_cell.font = font_bold
        title_cell.alignment = align_center
        ws.row_dimensions[excel_row_index].height = 25

        # Move down to write data
        excel_row_index += 1

        # Write data to rows
        for rfp in rfp_list:
            hyperlink_formula = f"=HYPERLINK(I{excel_row_index}, J{excel_row_index})"

            # --- LOGIC: DISCIPLINE COLORS ---
            # Default to None/Grey
            val_ns, fill_ns_cell = "", fill_data_row
            val_ss, fill_ss_cell = "", fill_data_row
            val_h, fill_h_cell = "", fill_data_row

            # Check Natural Sciences
            if rfp.get("Disc_NS", "").strip().upper() != "NONE":
                val_ns = rfp.get("Disc_NS", "Natural Sciences")
                fill_ns_cell = fill_ns  # Orange

            # Check Social Sciences
            if rfp.get("Disc_SS", "").strip().upper() != "NONE":
                val_ss = rfp.get("Disc_SS", "Social Sciences")
                fill_ss_cell = fill_ss  # Blue

            # Check Humanities
            if rfp.get("Disc_H", "").strip().upper() != "NONE":
                val_h = rfp.get("Disc_H", "Humanities")
                fill_h_cell = fill_h  # Green

            row_data = [
                rfp["Deadline"],
                rfp["Granting Organization"],
                hyperlink_formula,
                rfp["Funding Amount"],
                rfp["Keywords"],
                val_ns,
                val_ss,
                val_h,
                rfp["URL"],
                rfp["Grant Name"],
                " ",
            ]

            for col_num, value in enumerate(row_data, start=1):
                # This line is to make sure we increment our counter correctly to keep track of where we are in the worksheet
                cell = ws.cell(row=excel_row_index, column=col_num, value=value)
                cell.alignment = align_general

                if col_num == 11:
                    # force column k white for a cleaner look
                    cell.fill = fill_white
                else:
                    cell.fill = fill_data_row

                # Special Fills (Disciplines)
                if col_num == 6:  # NS
                    cell.fill = fill_ns_cell
                elif col_num == 7:  # SS
                    cell.fill = fill_ss_cell
                elif col_num == 8:  # H
                    cell.fill = fill_h_cell

                # Applies "link styling"
                if col_num == 3:
                    cell.font = font_link

            ws.row_dimensions[excel_row_index].height = 25
            excel_row_index += 1

        # Add spacing between blocks
        excel_row_index += 2

    # Now we loop through each of the month buckets, and write their contents using the block writing function we defined earlier
    sorted_keys = sorted([k for k in buckets.keys() if isinstance(k, date)])

    for key in sorted_keys:
        rfps_in_month = buckets[key]
        if rfps_in_month:
            # Only write if there is data (no empty blocks)
            month_title = key.strftime("%B %Y")
            write_rfp_list(rfps_in_month, month_title)

    # The rolling bucket gets added last
    if buckets["Rolling"]:
        write_rfp_list(buckets["Rolling"], "Rolling Deadlines")

    ## -------------------- COMPLETE - PRINT INFORMATION --------------------
    try:
        wb.save(full_path)
        print(f"Success! Excel Report generated at: {full_path}")
        print(
            f"Filter applied: Only dates between {start_date} and {end_date} (plus 'rolling')."
        )
    except PermissionError:
        print(f"Error: Could not save to '{full_path}'. Check if the file is open")
    except Exception as error:
        print(f"An unexpected error occurred: {error}")


def add_discipline_data_from_keywords(rfp_library):
    ## Scans the keywords field of the library, and appends with discipline data

    # if it includes one of these, it is definitly [discipline].
    ns_keywords = {
        "Climate & Environment",
        "Data Science",
        "Food & Agriculture",
        "Life Sciences/Health",
        "AI/ Machine Learning",
        "Neuroscience",
        "Physical Sciences",
        "Science/Technology",
        "Public Health",
    }

    ss_keywords = {
        "Business & Economics",
        "Law",
        "Economics",
        "Journalism",
        "Justice",
        "Policy",
        "Social Sciences",
    }

    h_keywords = {"Arts and Humanities", "Museums & Libraries"}

    for rfp in rfp_library:
        # Get the keywords string, defaulting to empty if missing
        raw_keywords = rfp.get("Keywords", "")

        # Split by comma and strip whitespace (Ex: "Data Science,Policy" -> ["Data Science", "Policy"])
        current_tags = [tag.strip() for tag in raw_keywords.split(",")]

        # Initialize flags
        is_ns = False
        is_ss = False
        is_h = False

        # Check against our lists
        for tag in current_tags:
            if tag in ns_keywords:
                is_ns = True
            if tag in ss_keywords:
                is_ss = True
            if tag in h_keywords:
                is_h = True

        # If it doesn't cleanly fit into one or more categories, assume it can apply to any
        if not (is_ns or is_ss or is_h):
            is_ns = True
            is_ss = True
            is_h = True

        # Make sure this dictionay outputs what the report maker function expects to find
        rfp["Disc_NS"] = "Natural Sciences" if is_ns else "NONE"
        rfp["Disc_SS"] = "Social Sciences" if is_ss else "NONE"
        rfp["Disc_H"] = "Humanities" if is_h else "NONE"

    print("Disciplines added based on provided keywords.")
    return rfp_library


def web_scraper_stanford(url):
    ## ---------- FUNCTION OBJECTIVES ------------
    # This function will scrape the website provided by "url" and return the following information:
    # Granting Organization
    # Grant Name
    # Funding Amount (May be N/A or Varies; if this is the case always return "Varies")
    # Deadline
    # Link (the hyperlink attached the text "guidelines")
    # (TBD): Keyword analysis from abstract

    ## -------------------- PHASE 1: SCRAPE HTML TEXT AND DEFINE BLOCKS --------------------

    print(f"fetching data from {url}")

    # define an empty library to add values to
    rfp_library = []

    try:
        # this will let the website know we are a person from a real computer, not a bot
        # this "user-agent" is accosiated with Blake Allen and his computer in ARM 172L.
        # if you are having problems with this code, try changing the user-agent to your own computer's
        # you can find this by googling "what is my user agent", and google will return it for you directly
        user_agent_blake = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:140.0) Gecko/20100101 Firefox/140.0"
        }

        response = requests.get(url, headers=user_agent_blake, timeout=15)

        # Check for errors loadig the webpage
        # 404 is the status code for page not found, 200 means it loadded correctly, etc
        if response.status_code != 200:
            print(f"Error: Failed to load page (Status Code: {response.status_code})")
            # if error occurs, print above message and return a blank library
            return []

    except Exception as e:
        print(f"Connection Error: {e}")
        # if error occurs, print above message and return a blank library
        return []

    # We don't want to include any of the ones that have a dealine passed,
    # which are all at the end of the page after the specific text "Deadline passed - for reference only"
    current_offers_only = response.text.split("Deadline passed - for reference only")[0]

    # turn the html text into more legible text
    soup = BeautifulSoup(current_offers_only, "html.parser")

    # This attempts to skip to the main portion of the text (main), and if that fails it just continues with the whole text file
    main_content = soup.find("main") or soup.body

    # Check to make sure we found the main content
    if not main_content:
        print("Error: Could not find main content area")
        # if error occurs, print above message and return a blank library
        return []

    # Get all text blocks (paragraphs and lists)
    org_names = [h for h in main_content.find_all("h3") if h.get_text(strip=True)]

    # generate an empty list of blocks to be filled
    grant_blocks = []

    for header in org_names:
        # This list will hold every HTML tag (p, div, ul) between this header and the next one
        block_content = []

        # Start looking at the element immediately after the current <h3> header
        current_element = header.find_next_sibling()

        # Keep grabbing elements until we hit the next <h3> header or run out of content
        while current_element:
            # If we hit an <h3>, that means we've reached the NEXT organization. Stop.
            if current_element.name == "h3":
                break

            # Add the element to our current block
            block_content.append(current_element)

            # Move to the next element
            current_element = current_element.find_next_sibling()

        # Append to the list and continue the loop for all organizaion names
        grant_blocks.append(
            {
                "organization_name": header.get_text(strip=True),
                "elements": block_content,
            }
        )

    ## -------------------- PHASE 2: FIND OTHER INFO IN EACH BLOCK --------------------

    # Get total count for the progress tracker
    total_grants = len(grant_blocks)

    # Use 'enumerate' to get the current number (i) automatically
    for i, block in enumerate(grant_blocks, 1):

        # Print the progress (end='\r' keeps it on the same line)
        print(f"Completed {i}/{total_grants} AI Keyword Analysis...", end="\r")

        # Start the entry with the Organization Name we already found
        entry = {
            "Granting Organization": block["organization_name"],
            "Grant Name": "General Research Grant",  # Default fallback
            # We will fill these in later steps:
            "Funding Amount": "Varies",
            "Deadline": "Rolling",
            "URL": "",
            "Keywords": "",
        }

        # ------ GRANT NAME ------
        # Find the FIRST <p> tag that contains a <strong> tag
        for element in block["elements"]:
            if element.name == "p":
                bold_tag = element.find("strong")

                # Once we found a bold tag, that's our grant name
                if bold_tag:
                    entry["Grant Name"] = bold_tag.get_text(strip=True)
                    # Since the grant name will ALWAYS be the first bolded text directly following the org_name, stop after 1
                    break
        # ------ /GRANT NAME ------

        # ------ DEADLINE ------
        for element in block["elements"]:
            text = element.get_text(separator=" ", strip=True)

            # Find the line starting with "Deadline:"
            if "Deadline:" in text:
                # Use Regex to grab the date part.
                # This grabs everything after "Deadline:" until a newline or open parenthesis (
                match = re.search(
                    r"Deadline:\s*([A-Za-z]+\s+\d{1,2},\s+\d{4})", text, re.IGNORECASE
                )

                if match:
                    raw_date_str = match.group(1).strip()  # e.g. "January 7, 2026"

                    # Try to convert "January 7, 2026" -> "01/07/2026"
                    try:
                        dt_object = datetime.strptime(raw_date_str, "%B %d, %Y")
                        entry["Deadline"] = dt_object.strftime("%m/%d/%Y")
                    except ValueError:
                        # If the format is weird, just keep the raw text (or default to Rolling)
                        # But usually, it's safer to default to "Rolling" so the Report Maker buckets it correctly
                        entry["Deadline"] = "Rolling"

                # Stop looking after we find it, since there should not be multiple deadlines in one block
                break
        # ------ /DEADLINE ------

        # ------ FUNDING AMOUNT ------
        for element in block["elements"]:
            text = element.get_text(separator=" ", strip=True)

            # Case-insensitive check for the label
            if "Funding amount:" in text:
                # Split the text at "Funding amount:" and take the second part (the actual data)
                # We use regex split to be safe against "Funding Amount:" vs "Funding amount:"
                parts = re.split(r"Funding amount:", text, flags=re.IGNORECASE)

                if len(parts) > 1:
                    full_amount_text = parts[1].strip()

                    # Logic: If it is just "N/A" or "Varies", keep it simple.
                    # Otherwise, keep the FULL robust description.
                    if full_amount_text.lower() in ["n/a", "varies"]:
                        entry["Funding Amount"] = "Varies"
                    else:
                        entry["Funding Amount"] = full_amount_text

                # Stop looking after the first instance
                break
        # ------ /FUNDING AMOUNT ------

        # ------ URL ------
        # in case we don't find anything, we want the program to keep running
        backup_link = ""

        for element in block["elements"]:
            # Skip elements that aren't tags (like empty text space)
            if not hasattr(element, "find_all"):
                continue

            # Find all links in this paragraph/element
            links = element.find_all("a", href=True)

            for link in links:
                link_text = link.get_text(strip=True).lower()
                href = link["href"]

                # 1. Save the very first link we find as a backup
                if not backup_link:
                    backup_link = href

                # 2. Check if this link is the "Gold Standard" (contains "guidelines")
                if "guidelines" in link_text:
                    entry["URL"] = href
                    break  # We found the best one, stop looking in this paragraph

            # If we found the specific "guidelines" link, stop looking through the rest of the block
            if entry["URL"]:
                break
        # Final Check: If we never found a "guidelines" link, use the backup
        if not entry["URL"]:
            entry["URL"] = backup_link
        # ------ /URL ------

        # ------ KEYWORDS ------
        description_parts = []

        # Add the Grant Name first as it is high-value context
        description_parts.append(entry["Grant Name"])

        for element in block["elements"]:
            # Get plain text (strip=True removes HTML tags)
            text = element.get_text(separator=" ", strip=True)

            # Filter out the lines we already scraped (not relevant to keywords)
            if "Funding amount:" in text:
                continue
            if "Deadline:" in text:
                continue
            if "guidelines" in text.lower() and len(text) < 100:
                # Heuristic: if a line mentions guidelines and is short, it's likely just the link line
                continue

            description_parts.append(text)

        # Join into one big string to analyse
        clean_text = " ".join(description_parts)

        # Run the function
        entry["Keywords"] = get_keywords(clean_text)
        # ------ /KEYWORDS ------

        rfp_library.append(entry)

    # TEMP Print results to check
    # for item in rfp_library:
    # print(f"Org: {item['Granting Organization']}")
    # print(f"Grant: {item['Grant Name']}")
    # print(f"Funding Amount: {item['Funding Amount']}")
    # print(f"Deadline: {item['Deadline']}")
    # print(f"URL: {item['URL']}")
    # print(f"Keywords: {item['Keywords']}")
    # print("-" * 20)

    ## ------------ FINISHED ------------

    ## Now that we have iterated through them all, we can export the final list to be used in the report generator
    print(f"Extracted {len(rfp_library)} grants from {url}!")
    return rfp_library


def analyze_local_html(filename):
    # just to remove an error message later on
    timezones = {
        "ET": 0,
        "EST": 0,
        "EDT": 0,
        "CT": 0,
        "CST": 0,
        "CDT": 0,
        "PT": 0,
        "PST": 0,
        "PDT": 0,
    }
    print(f"Opening '{filename}' for analysis...")

    ## -------------------- PHASE 1: SCRAPE HTML TEXT AND DEFINE BLOCKS --------------------

    # Read the contents to be analyzed with soup
    with open(filename, "r", encoding="utf-8") as f:
        html_content = f.read()

    # This turns it into more readable data for the computer
    soup = BeautifulSoup(html_content, "html.parser")
    # define an empty library to add values to
    rfp_library = []

    # Find the H2 with the specific ID, then get its parent div
    # This effectively finds the <div class="cards-title-row">...</div> block
    start_marker = soup.find("h2", id="search-result-count").parent
    end_marker = soup.find("div", id="paginationContainer")

    main_content = []
    # This will find everything inbetween the start and end markers
    for sibling in start_marker.find_next_siblings():
        if sibling == end_marker:
            break
        main_content.append(sibling)

    # Finds all h3 headers with the specific class "card-title", wihch on this site is all grant names
    # We have to check the entire soup, not just main content, becuase soup is a tree and main_content is a string
    all_h3_headers = soup.find_all("h3", class_="card-title")

    orgs_and_names = []

    for header in all_h3_headers:
        # Get the full text: "[Org Name]: [Fund Name]"
        full_text = header.get_text(strip=True)

        # Split the string at the 1st colon
        parts = full_text.split(":", 1)

        if len(parts) == 2:
            org_name = parts[0].strip()
            grant_name = parts[1].strip()
        else:
            # Handle cases where there might not be a colon
            org_name = full_text
            grant_name = full_text

        orgs_and_names.append({"Organization": org_name, "Grant": grant_name})

    rfp_library = []
    i = 0

    for header in all_h3_headers:
        i = i + 1
        print(f"Completed {i}/{len(all_h3_headers)} AI Keyword Analysis...", end="\r")

        # ------ GRANT ORG + NAME ------
        full_title = header.get_text(strip=True)
        parts = full_title.split(":", 1)
        org_name = parts[0].strip() if len(parts) > 1 else full_title
        grant_name = parts[1].strip() if len(parts) > 1 else full_title
        # ------ /GRANT ORG + NAME ------

        # grabs all text inside the card with that h3 title
        card_body = header.parent

        # ----- KEYWORDS ------
        # Find the paragraph tag with class 'card-text' inside this specific card body
        description_tag = card_body.find("p", class_="card-text")
        # Did we find it?
        description_text = (
            description_tag.get_text(strip=True)
            if description_tag
            else "no desciption found"
        )
        # Call an AI model to get keywords
        keywords = get_keywords(description_text)
        # ----- /KEYWORDS ------

        # ------ FUNDING & DEADLINE ------
        # These are key-value pairs. A robust way is to find the label, then grab the value next to it
        funding_amt = "N/A"
        deadline = "rolling"

        # Find all "keys" (labels) in this card
        all_keys = card_body.find_all("div", class_="category-key")

        for key_element in all_keys:
            key_text = key_element.get_text(strip=True)

            # Check if this key is one we want
            if "Funding Amt" in key_text:
                # The value is the next sibling div
                val_element = key_element.find_next_sibling(
                    "div", class_="category-value"
                )
                if val_element:
                    funding_amt = val_element.get_text(strip=True)

            elif "Due Date" in key_text:
                val_element = key_element.find_next_sibling(
                    "div", class_="category-value"
                )
                if val_element:
                    deadline = val_element.get_text(strip=True)

        if deadline and deadline not in ["rolling"]:
            try:
                # 2. Pass the 'tzinfos' argument here to silence the warning
                dt = parser.parse(deadline, fuzzy=True, tzinfos=timezones)

                # This strips the time/timezone info entirely, leaving just the date
                deadline = dt.strftime("%m/%d/%Y")
            except:
                pass
        # ------ /FUNDING & DEADLINE ------

        # ------ URL ------
        # The URL is in the footer, which is OUTSIDE the card_body
        # It is usually the next sibling of the card_body
        url = "N/A"
        card_footer = card_body.find_next_sibling("div", class_="card-footer")

        if card_footer:
            link_tag = card_footer.find(
                "a", string=lambda text: text and "Learn More" in text
            )
            if link_tag:
                raw_url = link_tag.get("href")
                # Fix relative links
                if raw_url.startswith("#"):
                    url = "https://princeton.infoready4.com/" + raw_url
                else:
                    url = raw_url
        # ------ /URL ------

        # Append the entries we found
        rfp_library.append(
            {
                "Granting Organization": org_name,
                "Grant Name": grant_name,
                "Funding Amount": funding_amt,
                "Deadline": deadline,
                "URL": url,
                "Keywords": keywords,
            }
        )
    print(f"Found {len(rfp_library)} sources from Princeton!")
    return rfp_library


def web_scraper_princeton(url):
    html_text_file = asyncio.run(async_princeton(url))
    rfp_library = analyze_local_html(html_text_file)
    return rfp_library
